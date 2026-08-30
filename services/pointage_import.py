"""Import du canevas de pointage mensuel (fichier canevas_<Mois>_AAAA.xlsx).

Structure reconnue (feuille nommee « MM-AAAA », ex. « 08-2026 ») :
  - ligne 1  : dates des jours (colonnes N, P, R, ... une paire par jour)
  - ligne 12 : en-tetes (Matricule, Nom, CIN, Fonction, Situation, ...)
  - ligne 13 et suivantes : un ouvrier par ligne
  - colonnes de pointage : Presence (paire) + H.Supp (impaire) pour chaque jour

Le detail quotidien n'est pas stocke : on calcule les agregats du mois
(jours travailles, jours ouvres, heures supp, taux de presence).
"""
import os
import re
from datetime import date, datetime

import openpyxl

from models.db import db
from models.mainoeuvre import Ouvrier
from models.metier import JournalImport, Parametre
from services.contexte import projet_actif_id


def _txt(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def _presence(v):
    """Convertit une cellule de pointage en fraction de journee (1 / 0,5 / 0).

    Les codes texte (AT, ML, SO, « Soldé », « Demande STC ») comptent comme
    absence (0). Les grands nombres sont des heures supplementaires mal placees
    et sont ignores ici (traites separement)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        if f in (0.0, 0.5, 1.0):
            return f
        return 0.0  # valeur numerique hors 0/0,5/1 : traitee comme non-presence
    t = _txt(v).replace(",", ".")
    try:
        f = float(t)
        return f if f in (0.0, 0.5, 1.0) else 0.0
    except ValueError:
        return 0.0  # AT, ML, SO, Soldé, Demande STC...


def _heure_supp(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(_txt(v).replace(",", "."))
    except ValueError:
        return 0.0


def _mois_depuis_feuille(nom_feuille):
    """« 08-2026 » -> « 2026-08 ». Renvoie None si le nom ne correspond pas."""
    m = re.match(r"^\s*(\d{1,2})-(\d{4})\s*$", nom_feuille)
    if m:
        return f"{m.group(2)}-{int(m.group(1)):02d}"
    return None


def _trouver_feuille_pointage(wb):
    """Retourne (feuille, mois) : la premiere feuille au format MM-AAAA."""
    for nom in wb.sheetnames:
        mois = _mois_depuis_feuille(nom)
        if mois:
            return wb[nom], mois
    return None, None


def analyser(chemin):
    """Apercu avant import : feuille de pointage detectee, mois, nb d'ouvriers."""
    wb = openpyxl.load_workbook(chemin, data_only=True)
    ws, mois = _trouver_feuille_pointage(wb)
    if ws is None:
        wb.close()
        return {"reconnu": False, "feuilles": wb.sheetnames}

    nb = 0
    for r in range(13, ws.max_row + 1):
        if _txt(ws.cell(r, 4).value):
            nb += 1
    # colonnes de jours (ligne 1, une date par paire)
    jours = sum(1 for c in range(14, ws.max_column + 1) if ws.cell(1, c).value is not None)
    wb.close()
    return {"reconnu": True, "feuille": ws.title, "mois": mois, "ouvriers": nb, "jours": jours}


def importer(chemin, utilisateur="systeme", projet_id=None):
    """Import transactionnel d'un canevas mensuel dans le projet actif.

    Remplace les ouvriers du meme mois pour ce projet (reimport possible),
    conserve les autres mois et les autres projets. En cas d'erreur : rollback
    complet, donnees existantes preservees.
    """
    pid = projet_id or projet_actif_id()
    try:
        wb = openpyxl.load_workbook(chemin, data_only=True)
        ws, mois = _trouver_feuille_pointage(wb)
        if ws is None:
            wb.close()
            return False, {"erreur": "Aucune feuille de pointage au format MM-AAAA."}

        # Colonnes de presence : paires a partir de N (14), pas de 2
        cols_presence = [c for c in range(14, ws.max_column + 1, 2) if ws.cell(1, c).value is not None]

        # On efface le mois concerne (pour ce projet) puis on reinsere (idempotent)
        Ouvrier.query.filter_by(mois=mois, projet_id=pid).delete()

        n = 0
        for r in range(13, ws.max_row + 1):
            nom = _txt(ws.cell(r, 4).value)
            if not nom:
                continue

            travailles = 0.0
            ouvres = 0
            supp = 0.0
            for c in cols_presence:
                p = _presence(ws.cell(r, c).value)
                if p is not None:
                    travailles += p
                    ouvres += 1
                supp += _heure_supp(ws.cell(r, c + 1).value)

            date_entree = ws.cell(r, 13).value
            if isinstance(date_entree, datetime):
                date_entree = date_entree.date()
            elif not isinstance(date_entree, date):
                date_entree = None

            db.session.add(
                Ouvrier(
                    projet_id=pid,
                    mois=mois,
                    matricule_chantier=_txt(ws.cell(r, 2).value),
                    matricule_sage=_txt(ws.cell(r, 3).value),
                    nom=nom,
                    cin=_txt(ws.cell(r, 5).value),
                    section=_txt(ws.cell(r, 6).value),
                    fonction=_txt(ws.cell(r, 7).value) or "NON RENSEIGNE",
                    situation=_txt(ws.cell(r, 8).value),
                    date_entree=date_entree,
                    jours_travailles=round(travailles, 2),
                    jours_ouvres=ouvres,
                    heures_supp=round(supp, 2),
                    taux_presence=round(100.0 * travailles / ouvres, 2) if ouvres else 0.0,
                )
            )
            n += 1

        wb.close()
        Parametre.set("dernier_import_pointage", datetime.utcnow().isoformat(timespec="seconds"), projet_id=pid)
        db.session.add(
            JournalImport(
                projet_id=pid,
                fichier=os.path.basename(chemin),
                utilisateur=utilisateur,
                statut="OK",
                resume=f"pointage {mois} : ouvriers={n}",
            )
        )
        db.session.commit()
        return True, {"mois": mois, "ouvriers": n}

    except Exception as exc:  # securite : on ne casse jamais la base
        db.session.rollback()
        db.session.add(
            JournalImport(
                projet_id=pid,
                fichier=os.path.basename(chemin),
                utilisateur=utilisateur,
                statut="ERREUR",
                resume=f"pointage : {str(exc)[:400]}",
            )
        )
        db.session.commit()
        return False, {"erreur": str(exc)}
