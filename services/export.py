"""Export des donnees filtrees en Excel (.xlsx) ou CSV."""
import csv
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .tables import TABLES, toutes_les_lignes

VERT_ENTETE = "1B6B43"


def _colonnes(spec):
    return [(c["cle"], c["label"]) for c in spec.colonnes]


def vers_csv(cle_table, params):
    spec = TABLES[cle_table]
    lignes = toutes_les_lignes(cle_table, params)
    cols = _colonnes(spec)
    tampon = io.StringIO()
    ecrivain = csv.writer(tampon, delimiter=";")
    ecrivain.writerow([lbl for _, lbl in cols])
    for l in lignes:
        ecrivain.writerow([l.get(c, "") for c, _ in cols])
    return tampon.getvalue().encode("utf-8-sig")


def vers_excel(cle_table, params):
    spec = TABLES[cle_table]
    lignes = toutes_les_lignes(cle_table, params)
    cols = _colonnes(spec)

    wb = Workbook()
    ws = wb.active
    ws.title = spec.titre[:31]

    ws.append([lbl for _, lbl in cols])
    entete_fill = PatternFill("solid", fgColor=VERT_ENTETE)
    for cellule in ws[1]:
        cellule.font = Font(bold=True, color="FFFFFF")
        cellule.fill = entete_fill
        cellule.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for l in lignes:
        ws.append([l.get(c, "") for c, _ in cols])

    for i, (cle, lbl) in enumerate(cols, start=1):
        largeur = max(len(lbl) + 4, 14)
        ws.column_dimensions[get_column_letter(i)].width = largeur

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"

    flux = io.BytesIO()
    wb.save(flux)
    flux.seek(0)
    return flux.read()


def nom_fichier(cle_table, extension):
    horodatage = datetime.now().strftime("%Y%m%d_%H%M")
    return f"CASAONE_{cle_table}_{horodatage}.{extension}"
