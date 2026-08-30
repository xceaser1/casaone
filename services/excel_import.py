"""Lecture, analyse et import du classeur Excel CASA ONE.

Le classeur n'est pas une base de donnees : ce sont des tableaux croises avec
des en-tetes fusionnes, des colonnes en paires (surface prevue / surface coulee)
et des libelles saisis a la main. Ce module fait tout le travail de nettoyage
puis ecrit dans SQLite au format long.

Regle d'or : on ne modifie JAMAIS le fichier Excel source, on le lit seulement.
"""
import os
import re
from datetime import date, datetime

import openpyxl

from models.db import db
from models.metier import (
    Betonnage,
    Bloc,
    DalleSurface,
    Decompte,
    JournalImport,
    Niveau,
    Parametre,
    Surface,
    ValidationPlan,
    Zone,
)
from services.contexte import projet_actif_id

# --------------------------------------------------------------------------
# Referentiels
# --------------------------------------------------------------------------

# Ordre physique des niveaux (du bas vers le haut)
NIVEAUX = [
    ("DALL", "Dallage", 0),
    ("PHSS3", "Plancher haut sous-sol 3", 1),
    ("PHSS2", "Plancher haut sous-sol 2", 2),
    ("PHMSS1", "Plancher haut mezzanine SS1", 3),
    ("PHSS1", "Plancher haut sous-sol 1", 4),
    ("PHRDC", "Plancher haut RDC", 5),
    ("PH ETG 1", "Plancher haut etage 1", 6),
    ("PH ETG 2", "Plancher haut etage 2", 7),
    ("PH ETG 3", "Plancher haut etage 3", 8),
    ("PH ETG 4", "Plancher haut etage 4", 9),
    ("PH ETG 5", "Plancher haut etage 5", 10),
    ("PH ETG 6", "Plancher haut etage 6", 11),
    ("PH ETG 7", "Plancher haut etage 7", 12),
    ("PH ETG 8", "Plancher haut etage 8", 13),
]

# Libelles libres rencontres dans la feuille Betonnage -> code niveau normalise
ALIAS_NIVEAUX = {
    "DALLAGE": "DALL",
    "DALL": "DALL",
    "RDC": "PHRDC",
    "PHRDC": "PHRDC",
    "PH RDC": "PHRDC",
    "PH1ER ETAGE": "PH ETG 1",
    "PH 1ER ETAGE": "PH ETG 1",
    "PH2EME ETAGE": "PH ETG 2",
    "PH3EME ETAGE": "PH ETG 3",
    "PH4EME ETAGE": "PH ETG 4",
    "PH5EME ETAGE": "PH ETG 5",
    "PH6EME ETAGE": "PH ETG 6",
    "PH7EME ETAGE": "PH ETG 7",
    "PH8EME ETAGE": "PH ETG 8",
}
for i in range(1, 9):
    ALIAS_NIVEAUX[f"PH ETG {i}"] = f"PH ETG {i}"
    ALIAS_NIVEAUX[f"PHETG{i}"] = f"PH ETG {i}"

# Les codes du referentiel sont evidemment leurs propres alias (PHSS3, PHSS1, ...)
for _code, _lib, _ordre in NIVEAUX:
    ALIAS_NIVEAUX.setdefault(_code, _code)

BLOC_LIBELLES = {
    "MAO": "Mail Ouest",
    "MAC": "Mail Central",
    "MAE": "Mail Est",
    "IMM1": "Immeuble 1",
    "IMMA": "Immeuble A",
    "IMMB": "Immeuble B",
    "HOTEL": "Hotel",
    "GBM": "Grande Boite Marche",
    "ESP": "Esplanade",
}

TYPES_DALLE = {
    "Dalle reticulee": "Dalle Reticulee",
    "Dalle Pleine": "Dalle Pleine",
    "Dalle Post-Tension": "Dalle Post-Tension",
    "Dalle Hourdis": "Dalle Hourdis",
}

FEUILLES_ATTENDUES = [
    "Tableau de surfaces",
    "Betonnage mensuelle",
    "VALIDATION DES PLANS",
    "Dalle reticulee",
    "Dalle Pleine",
    "Dalle Post-Tension",
    "Dalle Hourdis",
    "SUIVIE DES COUTS",
]


# --------------------------------------------------------------------------
# Utilitaires
# --------------------------------------------------------------------------
def _txt(valeur):
    """Nettoie une cellule texte (retours a la ligne, espaces multiples)."""
    if valeur is None:
        return ""
    return re.sub(r"\s+", " ", str(valeur)).strip()


def _nombre(valeur):
    """Convertit une cellule en float, 0.0 si vide ou non numerique."""
    if valeur is None or isinstance(valeur, (datetime, date)):
        return 0.0
    if isinstance(valeur, (int, float)):
        return float(valeur)
    txt = str(valeur).replace("\u00a0", "").replace(" ", "").replace(",", ".")
    try:
        return float(txt)
    except ValueError:
        return 0.0


def _sans_accents(texte):
    table = str.maketrans("àâäéèêëîïôöùûüçÀÂÄÉÈÊËÎÏÔÖÙÛÜÇ", "aaaeeeeiioouuucAAAEEEEIIOOUUUC")
    return texte.translate(table)


def norme_niveau(valeur):
    """Ramene un libelle de niveau saisi a la main vers un code du referentiel."""
    brut = _sans_accents(_txt(valeur)).upper()
    if not brut:
        return None
    if brut in ALIAS_NIVEAUX:
        return ALIAS_NIVEAUX[brut]
    compact = brut.replace(" ", "")
    for alias, code in ALIAS_NIVEAUX.items():
        if alias.replace(" ", "") == compact:
            return code
    m = re.match(r"^PH(\d+)(ER|EME)?ETAGE$", compact)
    if m:
        return f"PH ETG {int(m.group(1))}"
    return None


def norme_zone(valeur):
    """Extrait le code zone d'un libelle libre : "C7 95,90" -> "C7", "H2" -> "H"."""
    brut = _txt(valeur).upper()
    if not brut:
        return None
    premier = re.split(r"[;/+]", brut)[0].strip()
    m = re.match(r"^([ABC])\s*(\d{1,2})", premier)
    if m:
        return f"{m.group(1)}{int(m.group(2))}"
    if premier.startswith("H"):
        return "H"
    return None


def norme_statut(valeur):
    """Uniformise les statuts de plans ("En  Cours", "EN Cours" -> "En Cours")."""
    brut = _sans_accents(_txt(valeur)).upper()
    if not brut:
        return None
    if brut.startswith("NON"):
        return "Non Valide"
    if brut.startswith("VALID"):
        return "Valide"
    if "COURS" in brut:
        return "En Cours"
    return "En Cours"


def _ouvrir(chemin):
    return openpyxl.load_workbook(chemin, data_only=True, read_only=False)


def _feuille(wb, nom_recherche):
    """Retrouve une feuille meme si son nom comporte des espaces/accents differents."""
    cible = _sans_accents(nom_recherche).upper().strip()
    for nom in wb.sheetnames:
        if _sans_accents(nom).upper().strip() == cible:
            return wb[nom]
    return None


# --------------------------------------------------------------------------
# Analyse (aper�u avant import)
# --------------------------------------------------------------------------
def analyser(chemin):
    """Renvoie un descriptif de chaque feuille : lignes reelles, colonnes, apercu."""
    wb = _ouvrir(chemin)
    rapport = []
    for nom in wb.sheetnames:
        ws = wb[nom]
        lignes_reelles = 0
        apercu = []
        for ligne in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200), values_only=True):
            if any(c is not None and _txt(c) != "" for c in ligne):
                lignes_reelles += 1
                if len(apercu) < 6:
                    apercu.append([_txt(c) for c in ligne[:14]])
        rapport.append(
            {
                "feuille": nom,
                "lignes_declarees": ws.max_row,
                "colonnes": ws.max_column,
                "lignes_non_vides": lignes_reelles,
                "vide": lignes_reelles == 0,
                "reconnue": any(
                    _sans_accents(nom).upper().strip() == _sans_accents(f).upper().strip()
                    for f in FEUILLES_ATTENDUES
                ),
                "apercu": apercu,
            }
        )
    wb.close()
    return rapport


# --------------------------------------------------------------------------
# Referentiels : niveaux, blocs, zones
# --------------------------------------------------------------------------
def _assurer_niveaux(pid):
    cache = {n.code: n for n in Niveau.query.filter_by(projet_id=pid).all()}
    for code, libelle, ordre in NIVEAUX:
        if code not in cache:
            n = Niveau(code=code, libelle=libelle, ordre=ordre, projet_id=pid)
            db.session.add(n)
            cache[code] = n
    db.session.flush()
    return cache


def _assurer_blocs_zones(wb, pid):
    """Le couple (bloc, zone) est lu dans la feuille "Tableau de surfaces"."""
    ws = _feuille(wb, "Tableau de surfaces")
    cache_blocs = {b.code: b for b in Bloc.query.filter_by(projet_id=pid).all()}
    cache_zones = {z.code: z for z in Zone.query.filter_by(projet_id=pid).all()}
    ordre = 0
    for r in range(6, 40):
        bloc_brut = _txt(ws.cell(r, 4).value)
        zone_brut = _txt(ws.cell(r, 5).value)
        if not zone_brut or bloc_brut.upper().startswith(("SURFACE", "TAUX")):
            continue
        code_bloc = _sans_accents(bloc_brut).upper().split(" ")[0].strip()
        code_bloc = re.sub(r"[^A-Z0-9]", "", code_bloc) or "NC"
        if code_bloc == "IMM1":
            code_bloc = "IMM1"
        code_zone = norme_zone(zone_brut) or zone_brut.upper()
        if code_bloc not in cache_blocs:
            b = Bloc(code=code_bloc, libelle=BLOC_LIBELLES.get(code_bloc, code_bloc), projet_id=pid)
            db.session.add(b)
            db.session.flush()
            cache_blocs[code_bloc] = b
        if code_zone not in cache_zones:
            ordre += 1
            z = Zone(code=code_zone, bloc_id=cache_blocs[code_bloc].id, ordre=ordre, projet_id=pid)
            db.session.add(z)
            cache_zones[code_zone] = z
    db.session.flush()
    return cache_blocs, cache_zones


# --------------------------------------------------------------------------
# Feuilles en tableau croise (surfaces / validation / dalles)
# --------------------------------------------------------------------------
def _paires_niveaux(ws, ligne_entete, col_debut):
    """Repere les paires de colonnes (niveau, colonne de valeur coulee).

    En-tete type : DALL | DC | PHSS3 | PC | PHSS2 | PC | ...
    -> [(col_niveau, col_secondaire, code_niveau), ...]
    """
    paires = []
    col = col_debut
    while col <= ws.max_column:
        code = norme_niveau(ws.cell(ligne_entete, col).value)
        if code:
            suivant = _txt(ws.cell(ligne_entete, col + 1).value).upper()
            col_sec = col + 1 if suivant in ("DC", "PC", "SC") else None
            paires.append((col, col_sec, code))
            col += 2 if col_sec else 1
        else:
            col += 1
    return paires


def _importer_surfaces(wb, niveaux, zones, pid):
    ws = _feuille(wb, "Tableau de surfaces")
    paires = _paires_niveaux(ws, 5, 6)
    Surface.query.filter_by(projet_id=pid).delete()
    n = 0
    for r in range(6, 40):
        code_zone = norme_zone(ws.cell(r, 5).value)
        libelle = _txt(ws.cell(r, 4).value).upper()
        if not code_zone or code_zone not in zones or libelle.startswith(("SURFACE", "TAUX")):
            continue
        for col_tot, col_coul, code_niv in paires:
            total = _nombre(ws.cell(r, col_tot).value)
            coule = _nombre(ws.cell(r, col_coul).value) if col_coul else 0.0
            if total == 0 and coule == 0:
                continue
            db.session.add(
                Surface(
                    projet_id=pid,
                    zone_id=zones[code_zone].id,
                    niveau_id=niveaux[code_niv].id,
                    surface_totale=round(total, 2),
                    surface_coulee=round(coule, 2),
                )
            )
            n += 1
    return n


def _importer_validations(wb, niveaux, zones, pid):
    ws = _feuille(wb, "VALIDATION DES PLANS")
    # Ici les paires sont (surface, statut) : DALL | DC(statut) | PHSS3 | PC(statut) ...
    paires = _paires_niveaux(ws, 5, 4)
    ValidationPlan.query.filter_by(projet_id=pid).delete()
    n = 0
    for r in range(6, 40):
        code_zone = norme_zone(ws.cell(r, 3).value)
        libelle = _txt(ws.cell(r, 2).value).upper()
        if not code_zone or code_zone not in zones or libelle.startswith(("SURFACE", "TAUX")):
            continue
        for col_surf, col_stat, code_niv in paires:
            surface = _nombre(ws.cell(r, col_surf).value)
            statut = norme_statut(ws.cell(r, col_stat).value) if col_stat else None
            if surface == 0 and not statut:
                continue
            db.session.add(
                ValidationPlan(
                    projet_id=pid,
                    zone_id=zones[code_zone].id,
                    niveau_id=niveaux[code_niv].id,
                    surface=round(surface, 2),
                    statut=statut or "En Cours",
                )
            )
            n += 1
    return n


def _importer_dalles(wb, niveaux, zones, pid):
    DalleSurface.query.filter_by(projet_id=pid).delete()
    total = 0
    for nom_feuille, type_dalle in TYPES_DALLE.items():
        ws = _feuille(wb, nom_feuille)
        if ws is None:
            continue
        # La ligne d'en-tete est celle qui contient le libelle du type de dalle
        ligne_entete = None
        for r in range(1, 12):
            if _sans_accents(_txt(ws.cell(r, 1).value)).upper().startswith("DALLE"):
                ligne_entete = r
                break
        if ligne_entete is None:
            continue
        paires = _paires_niveaux(ws, ligne_entete, 3)
        for r in range(ligne_entete + 1, ligne_entete + 32):
            code_zone = norme_zone(ws.cell(r, 2).value)
            libelle = _txt(ws.cell(r, 1).value).upper()
            if not code_zone or code_zone not in zones or libelle.startswith("SURFACE"):
                continue
            for col_tot, col_coul, code_niv in paires:
                tot = _nombre(ws.cell(r, col_tot).value)
                coul = _nombre(ws.cell(r, col_coul).value) if col_coul else 0.0
                if tot == 0 and coul == 0:
                    continue
                db.session.add(
                    DalleSurface(
                        projet_id=pid,
                        type_dalle=type_dalle,
                        zone_id=zones[code_zone].id,
                        niveau_id=niveaux[code_niv].id,
                        surface_totale=round(tot, 2),
                        surface_coulee=round(coul, 2),
                    )
                )
                total += 1
    return total


# --------------------------------------------------------------------------
# Feuille "Betonnage mensuelle" : blocs mensuels de 5 mois cote a cote
# --------------------------------------------------------------------------
def _importer_betonnage(wb, niveaux, zones, pid):
    ws = _feuille(wb, "Betonnage mensuelle")
    Betonnage.query.filter_by(projet_id=pid).delete()
    # Les en-tetes de groupe sont les lignes contenant "JOUR" en colonne D
    lignes_entete = [
        r for r in range(1, ws.max_row + 1) if _txt(ws.cell(r, 4).value).upper() == "JOUR"
    ]
    colonnes_groupes = [4, 8, 12, 16, 20]  # D, H, L, P, T
    n = 0
    for idx, ligne_entete in enumerate(lignes_entete):
        fin = lignes_entete[idx + 1] - 3 if idx + 1 < len(lignes_entete) else ws.max_row
        for col in colonnes_groupes:
            date_courante = None
            for r in range(ligne_entete + 1, fin + 1):
                cel_date = ws.cell(r, col).value
                if isinstance(cel_date, (datetime, date)):
                    date_courante = cel_date.date() if isinstance(cel_date, datetime) else cel_date
                elif _txt(cel_date).upper().startswith("SURFACE"):
                    break  # ligne de total : fin du groupe
                bloc_brut = _txt(ws.cell(r, col + 1).value)
                niveau_brut = _txt(ws.cell(r, col + 2).value)
                surface = _nombre(ws.cell(r, col + 3).value)
                if not bloc_brut and not niveau_brut:
                    continue
                code_niv = norme_niveau(niveau_brut)
                code_zone = norme_zone(bloc_brut)
                if date_courante is None:
                    continue
                db.session.add(
                    Betonnage(
                        projet_id=pid,
                        date_coulage=date_courante,
                        zone_id=zones[code_zone].id if code_zone in zones else None,
                        niveau_id=niveaux[code_niv].id if code_niv in niveaux else None,
                        bloc_libelle=bloc_brut,
                        niveau_libelle=niveau_brut,
                        surface=round(surface, 2),
                        mois=date_courante.strftime("%Y-%m"),
                    )
                )
                n += 1
    return n


# --------------------------------------------------------------------------
# Feuille "SUIVIE DES COUTS"
# --------------------------------------------------------------------------
def _importer_couts(wb, pid):
    ws = _feuille(wb, "SUIVIE DES COUTS")
    if ws is None:
        return 0
    Decompte.query.filter_by(projet_id=pid).delete()
    n = 0
    ordre = 0
    for r in range(1, ws.max_row + 1):
        libelle = _txt(ws.cell(r, 1).value)
        montant = _nombre(ws.cell(r, 2).value)
        if not libelle or libelle.upper() in ("DECOMPTE", "DECOMPTE "):
            continue
        if _sans_accents(libelle).upper().startswith("DECOMPTE"):
            continue
        if libelle.upper() == "TOTAL":
            continue
        if montant <= 0:
            continue
        ordre += 1
        db.session.add(Decompte(projet_id=pid, libelle=libelle, montant=round(montant, 2), ordre=ordre))
        n += 1
    # Montant global du marche (cellule D18 dans le fichier d'origine)
    global_marche = 0.0
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            if "MONTANT GLOBAL" in _txt(ws.cell(r, c).value).upper():
                global_marche = _nombre(ws.cell(r + 1, c).value)
    if global_marche:
        Parametre.set("montant_global_marche", global_marche, "Montant global du marche (MAD)", projet_id=pid)
    return n


# --------------------------------------------------------------------------
# Point d'entree
# --------------------------------------------------------------------------
def importer(chemin, utilisateur="systeme", projet_id=None):
    """Import complet et transactionnel dans le projet actif (ou `projet_id`).

    En cas d'erreur, rollback : les donnees existantes sont conservees. Seules
    les donnees du projet cible sont remplacees ; les autres projets ne sont
    jamais touches.
    """
    pid = projet_id or projet_actif_id()
    resume = {}
    try:
        wb = _ouvrir(chemin)
        niveaux = _assurer_niveaux(pid)
        _, zones = _assurer_blocs_zones(wb, pid)
        resume["surfaces"] = _importer_surfaces(wb, niveaux, zones, pid)
        resume["validations"] = _importer_validations(wb, niveaux, zones, pid)
        resume["dalles"] = _importer_dalles(wb, niveaux, zones, pid)
        resume["betonnages"] = _importer_betonnage(wb, niveaux, zones, pid)
        resume["decomptes"] = _importer_couts(wb, pid)
        resume["zones"] = len(zones)
        Parametre.set("dernier_import", datetime.utcnow().isoformat(timespec="seconds"), projet_id=pid)
        Parametre.set("fichier_source", os.path.basename(chemin), projet_id=pid)
        db.session.add(
            JournalImport(
                projet_id=pid,
                fichier=os.path.basename(chemin),
                utilisateur=utilisateur,
                statut="OK",
                resume="; ".join(f"{k}={v}" for k, v in resume.items()),
            )
        )
        db.session.commit()
        wb.close()
        return True, resume
    except Exception as exc:  # pragma: no cover - securite
        db.session.rollback()
        db.session.add(
            JournalImport(
                projet_id=pid,
                fichier=os.path.basename(chemin),
                utilisateur=utilisateur,
                statut="ERREUR",
                resume=str(exc)[:2000],
            )
        )
        db.session.commit()
        return False, {"erreur": str(exc)}
