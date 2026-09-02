"""Gestion de stock : etat des depots, mouvements et transferts.

Toutes les ecritures passent par le journal des mouvements : une entree, une
sortie ou un transfert cree une ligne datee et nominative. Le stock affiche
est toujours recalcule a partir de ce journal.
"""
import csv
import io
from datetime import date

from flask import (Blueprint, Response, flash, redirect, render_template,
                   request, url_for)
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from models.db import db
from models.stock import TYPES, Article, Depot, Mouvement
from services import stock as svc
from services.contexte import projet_actif_id
from services.security import exige

bp = Blueprint("stock", __name__, url_prefix="/stock")


def _nombre(valeur, defaut=0.0):
    valeur = (valeur or "").strip().replace(" ", "").replace(",", ".")
    if not valeur:
        return defaut
    try:
        return float(valeur)
    except ValueError:
        return defaut


def _jour(valeur):
    try:
        return date.fromisoformat((valeur or "").strip())
    except ValueError:
        return date.today()


# --------------------------------------------------------------- Etat du stock
@bp.route("/")
@login_required
@exige("stock")
def index():
    pid = projet_actif_id()
    depots, lignes, alertes = svc.tableau(pid)
    return render_template(
        "stock.html", page="stock", depots=depots, lignes=lignes,
        alertes=alertes, filtres=svc.valeurs_filtres(pid), types=TYPES,
        aujourdhui=date.today(),
    )


# ------------------------------------------------------------- Mouvements
@bp.route("/mouvements")
@login_required
@exige("stock")
def mouvements():
    pid = projet_actif_id()
    q = Mouvement.query.filter_by(projet_id=pid)

    type_ = request.args.get("type")
    if type_ in dict(TYPES):
        q = q.filter(Mouvement.type == type_)
    if request.args.get("article"):
        q = q.filter(Mouvement.article_id == request.args.get("article", type=int))
    if request.args.get("depot"):
        d = request.args.get("depot", type=int)
        q = q.filter(db.or_(Mouvement.depot_source_id == d, Mouvement.depot_dest_id == d))
    if request.args.get("du"):
        q = q.filter(Mouvement.date_mouvement >= _jour(request.args["du"]))
    if request.args.get("au"):
        q = q.filter(Mouvement.date_mouvement <= _jour(request.args["au"]))

    lignes = q.order_by(Mouvement.date_mouvement.desc(), Mouvement.id.desc()).limit(500).all()
    return render_template(
        "stock_mouvements.html", page="stock-mouvements", lignes=lignes,
        filtres=svc.valeurs_filtres(pid), types=TYPES, args=request.args,
        aujourdhui=date.today(),
    )


@bp.route("/mouvements/nouveau", methods=["POST"])
@login_required
@exige("stock", "create")
def creer_mouvement():
    pid = projet_actif_id()
    type_ = request.form.get("type")
    if type_ not in dict(TYPES):
        flash("Type de mouvement inconnu.", "erreur")
        return redirect(url_for("stock.index"))

    article_id = request.form.get("article_id", type=int)
    quantite = _nombre(request.form.get("quantite"))
    source = request.form.get("depot_source_id", type=int)
    dest = request.form.get("depot_dest_id", type=int)

    # --- controles metier
    erreur = None
    if not article_id:
        erreur = "Choisissez un article."
    elif quantite <= 0:
        erreur = "La quantite doit etre superieure a zero."
    elif type_ == "entree" and not dest:
        erreur = "Choisissez le depot de destination."
    elif type_ == "sortie" and not source:
        erreur = "Choisissez le depot d'origine."
    elif type_ == "transfert":
        if not source or not dest:
            erreur = "Un transfert exige un depot d'origine et un depot de destination."
        elif source == dest:
            erreur = "Les deux depots d'un transfert doivent etre differents."

    # Le stock ne doit jamais devenir negatif : le magasinier regularise
    # d'abord l'entree correspondante.
    if erreur is None and type_ in ("sortie", "transfert"):
        dispo = svc.stock_article_depot(article_id, source, pid)
        if quantite > dispo + 1e-9:
            art = db.session.get(Article, article_id)
            unite = art.unite if art else ""
            erreur = f"Stock insuffisant : {dispo:g} {unite} disponible(s) dans ce depot."

    if erreur:
        flash(erreur, "erreur")
        return redirect(request.referrer or url_for("stock.index"))

    db.session.add(Mouvement(
        projet_id=pid, type=type_, article_id=article_id,
        depot_source_id=source if type_ in ("sortie", "transfert") else None,
        depot_dest_id=dest if type_ in ("entree", "transfert") else None,
        quantite=quantite, date_mouvement=_jour(request.form.get("date_mouvement")),
        reference=(request.form.get("reference") or "").strip(),
        motif=(request.form.get("motif") or "").strip(),
        saisi_par=current_user.username,
    ))
    db.session.commit()
    flash({"entree": "Entree enregistree.", "sortie": "Sortie enregistree.",
           "transfert": "Transfert enregistre."}[type_], "succes")
    return redirect(request.referrer or url_for("stock.index"))


@bp.route("/mouvements/<int:mid>/supprimer", methods=["POST"])
@login_required
@exige("stock", "delete")
def supprimer_mouvement(mid):
    m = Mouvement.query.filter_by(id=mid, projet_id=projet_actif_id()).first_or_404()
    db.session.delete(m)
    db.session.commit()
    flash("Mouvement supprime.", "succes")
    return redirect(request.referrer or url_for("stock.mouvements"))


# ------------------------------------------------------------------ Depots
@bp.route("/depots", methods=["POST"])
@login_required
@exige("stock", "create")
def creer_depot():
    pid = projet_actif_id()
    code = (request.form.get("code") or "").strip().upper()
    nom = (request.form.get("nom") or "").strip()
    if len(code) < 2 or not nom:
        flash("Code (2 caracteres minimum) et nom sont obligatoires.", "erreur")
    elif Depot.query.filter(Depot.projet_id == pid, db.func.upper(Depot.code) == code).first():
        flash("Ce code de depot existe deja.", "erreur")
    else:
        db.session.add(Depot(
            projet_id=pid, code=code, nom=nom,
            emplacement=(request.form.get("emplacement") or "").strip(),
            responsable=(request.form.get("responsable") or "").strip(),
        ))
        db.session.commit()
        flash("Depot cree.", "succes")
    return redirect(url_for("stock.index"))


# ---------------------------------------------------------------- Articles
@bp.route("/articles", methods=["POST"])
@login_required
@exige("stock", "create")
def creer_article():
    pid = projet_actif_id()
    code = (request.form.get("code") or "").strip().upper()
    designation = (request.form.get("designation") or "").strip()
    if not code or not designation:
        flash("Code et designation sont obligatoires.", "erreur")
    elif Article.query.filter(Article.projet_id == pid,
                              db.func.upper(Article.code) == code).first():
        flash("Ce code article existe deja.", "erreur")
    else:
        db.session.add(Article(
            projet_id=pid, code=code, designation=designation,
            categorie=(request.form.get("categorie") or "").strip(),
            unite=(request.form.get("unite") or "U").strip(),
            seuil_alerte=_nombre(request.form.get("seuil_alerte")),
        ))
        db.session.commit()
        flash("Article cree.", "succes")
    return redirect(url_for("stock.index"))


# ------------------------------------------------------------------ Export
@bp.route("/export.<fmt>")
@login_required
@exige("stock", "export")
def export(fmt):
    pid = projet_actif_id()
    depots, lignes, _ = svc.tableau(pid)
    entetes = ["Code", "Designation", "Categorie", "Unite"] + [d.code for d in depots] + ["Total"]

    def cellules(l):
        a = l["article"]
        return ([a.code, a.designation, a.categorie or "", a.unite]
                + [l["detail"].get(d.id, 0) for d in depots] + [l["total"]])

    nom = "stock_" + date.today().isoformat()

    if fmt == "csv":
        tampon = io.StringIO()
        w = csv.writer(tampon, delimiter=";")
        w.writerow(entetes)
        for l in lignes:
            w.writerow(cellules(l))
        return Response(
            tampon.getvalue().encode("utf-8-sig"),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="' + nom + '.csv"'})

    if fmt in ("xlsx", "excel"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock"
        ws.append(entetes)
        fill = PatternFill("solid", fgColor="1D4ED8")
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = fill
            c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for l in lignes:
            ws.append(cellules(l))
        flux = io.BytesIO()
        wb.save(flux)
        return Response(
            flux.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="' + nom + '.xlsx"'})

    return redirect(url_for("stock.index"))
