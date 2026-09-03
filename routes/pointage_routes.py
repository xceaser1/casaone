"""Pointage par badge QR : badges imprimables, scanner mobile, presences.

- /badges     : planche de badges QR (impression) — desktop.
- /scan       : scanner mobile (camera + QR) — ne fonctionne que sur telephone.
- /presences  : tableau des presences du jour — desktop.
- /api/checkin: enregistre une entree/sortie depuis un badge scanne ou un
                matricule saisi a la main.
"""
import csv
import io
import socket
from collections import Counter
from datetime import date, datetime, timedelta

from flask import Blueprint, Response, jsonify, render_template, request, url_for
from flask_login import current_user, login_required
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError, SQLAlchemyError

from models.db import db
from models.mainoeuvre import Ouvrier
from models.presence import Presence
from models.projet import Projet
from services import badges as svc_badges
from services.contexte import projet_actif_id
from services.security import exige

bp = Blueprint("pointage", __name__)

# Garde-fou sur la taille d'un lot de synchronisation.
MAX_LOT = 500


def _ip_locale():
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except OSError:
        return "127.0.0.1"


def _url_base():
    """URL racine accessible depuis le reseau (IP LAN + schema courant)."""
    port = request.host.split(":")[-1] if ":" in request.host else ("443" if request.scheme == "https" else "80")
    return f"{request.scheme}://{_ip_locale()}:{port}"


# Tolerance sur l'horloge du telephone : au-dela, on ne fait pas confiance a
# une heure situee dans le futur (telephone mal regle) et on retient l'heure
# du serveur.
DERIVE_HORLOGE = timedelta(minutes=5)


def heure_appareil(valeur):
    """Convertit une heure ISO envoyee par le telephone en datetime local.

    Renvoie None si la valeur est absente ou illisible : l'appelant retombe
    alors sur l'heure du serveur. Une heure dans le futur est ramenee a
    maintenant ; une heure passee est conservee telle quelle, car un pointage
    peut legitimement avoir ete enregistre hors ligne plusieurs jours plus tot.
    """
    if not valeur:
        return None
    try:
        h = datetime.fromisoformat(str(valeur).replace("Z", "+00:00"))
    except (TypeError, ValueError):
        return None
    if h.tzinfo is not None:
        h = h.astimezone().replace(tzinfo=None)
    maintenant = datetime.now()
    return maintenant if h > maintenant + DERIVE_HORLOGE else h


def _enregistrer(projet_id, matricule, nom, type_, methode, saisi_par,
                 heure=None, differe=False, commit=True):
    """Enregistre (ou met a jour) une presence. Renvoie (presence, deja_existant).

    `heure` permet de dater le pointage a l'heure de l'appareil : un badge
    scanne a 7 h 12 hors ligne et transmis a 18 h doit rester date de 7 h 12.
    Le jour est deduit de cette heure, pas de la date de reception.

    `differe` signale un pointage rejoue depuis la file d'attente locale. Dans
    ce cas on fusionne au lieu d'ecraser : une entree conserve l'heure la plus
    tot, une sortie la plus tard. Vider la file ne peut donc jamais degrader
    une heure deja correcte, quel que soit l'ordre d'arrivee.
    """
    if heure is None:
        heure = datetime.now()
    jour = heure.date()
    existant = Presence.query.filter_by(
        projet_id=projet_id, jour=jour, matricule=matricule, nom=nom, type=type_
    ).first()
    deja = existant is not None
    if existant:
        if differe:
            existant.heure = (
                min(existant.heure, heure) if type_ == "entree"
                else max(existant.heure, heure)
            )
        else:
            existant.heure = heure
            existant.methode = methode
            existant.saisi_par = saisi_par
        enreg = existant
    else:
        enreg = Presence(
            projet_id=projet_id, matricule=matricule, nom=nom, jour=jour, type=type_,
            heure=heure, methode=methode, saisi_par=saisi_par,
        )
        db.session.add(enreg)
    if commit:
        db.session.commit()
    return enreg, deja


def _resoudre(d):
    """Identifie l'ouvrier vise par un pointage (badge signe ou matricule).

    Renvoie (projet_id, matricule, nom, methode, erreur). L'erreur, quand elle
    est presente, est definitive : rejouer la meme donnee echouera pareil, donc
    le client doit la retirer de sa file plutot que de boucler indefiniment.
    """
    token = (d.get("token") or "").strip()
    if token:
        info = svc_badges.decoder_token(token)
        if info is None:
            return None, None, None, None, "Badge non reconnu."
        pid, mat, nom = info
        return pid, mat, nom, "scan", None

    pid = projet_actif_id()
    mat = (d.get("matricule") or "").strip()
    nom = (d.get("nom") or "").strip()
    if not (mat or nom):
        return None, None, None, None, "Matricule ou nom requis."
    if mat and not nom:
        row = (
            Ouvrier.query.filter_by(projet_id=pid, matricule_chantier=mat)
            .order_by(Ouvrier.mois.desc())
            .first()
        )
        if row is None:
            return None, None, None, None, f"Matricule « {mat} » introuvable."
        nom = row.nom or ""
    return pid, mat, nom, "manuel", None


def _fonction(projet_id, matricule, nom):
    """Fonction connue de l'ouvrier (dernier mois renseigne)."""
    q = Ouvrier.query.filter_by(projet_id=projet_id)
    q = q.filter_by(matricule_chantier=matricule) if matricule else q.filter_by(nom=nom)
    row = q.order_by(Ouvrier.mois.desc()).first()
    return (row.fonction or "") if row else ""


def _presences_jour(projet_id, jour):
    """Pivot par ouvrier : {matricule, nom, fonction, entree, sortie} pour un jour."""
    evenements = (
        Presence.query.filter_by(projet_id=projet_id, jour=jour)
        .order_by(Presence.heure.desc())
        .all()
    )
    par_ouvrier = {}
    for e in evenements:
        cle = (e.matricule, e.nom)
        d = par_ouvrier.setdefault(
            cle, {"matricule": e.matricule, "nom": e.nom, "fonction": "",
                  "entree": None, "sortie": None, "methode": e.methode}
        )
        if e.type == "entree" and d["entree"] is None:
            d["entree"] = e.heure
        elif e.type == "sortie":
            d["sortie"] = e.heure

    # Fonction de chaque ouvrier (dernier mois connu) : permet de filtrer par metier.
    if par_ouvrier:
        fonctions = {}
        for mat, nom, fonction in (
            db.session.query(Ouvrier.matricule_chantier, Ouvrier.nom, Ouvrier.fonction)
            .filter(Ouvrier.projet_id == projet_id)
            .order_by(Ouvrier.mois.asc())
            .all()
        ):
            if fonction:
                fonctions[(mat or "").strip()] = fonction
                fonctions[(nom or "").strip().upper()] = fonction
        for (mat, nom), d in par_ouvrier.items():
            d["fonction"] = (
                fonctions.get((mat or "").strip())
                or fonctions.get((nom or "").strip().upper())
                or ""
            )

    return sorted(par_ouvrier.values(), key=lambda x: (x["nom"] or "").upper())


# --------------------------------------------------------------------------
@bp.route("/badges")
@login_required
@exige("pointage")
def liste_badges():
    pid = projet_actif_id()
    # QR = URL publique de pointage : chaque ouvrier peut scanner son propre
    # badge avec l'appareil photo de son telephone.
    return render_template(
        "badges.html", page="badges", badges=svc_badges.badges(pid, base_url=_url_base())
    )


@bp.route("/scan")
@login_required
@exige("pointage")
def scan():
    # QR pointant vers cette page (avec l'IP reseau) : a scanner pour ouvrir le
    # scanner sur un telephone quand on arrive ici depuis un poste fixe.
    url_mobile = _url_base() + url_for("pointage.scan")
    return render_template(
        "scan.html",
        page="scan",
        url_mobile=url_mobile,
        qr_ouvrir=svc_badges.qr_svg(url_mobile, scale=6),
    )


@bp.route("/presences")
@login_required
@exige("pointage")
def presences():
    pid = projet_actif_id()
    valeur = request.args.get("jour")
    try:
        jour = date.fromisoformat(valeur) if valeur else date.today()
    except ValueError:
        jour = date.today()

    lignes = _presences_jour(pid, jour)

    # Effectif = roster du mois le plus recent importe (headcount attendu).
    dernier_mois = db.session.query(func.max(Ouvrier.mois)).filter(Ouvrier.projet_id == pid).scalar()
    effectif = (
        Ouvrier.query.filter_by(projet_id=pid, mois=dernier_mois).count() if dernier_mois else 0
    )

    # Fonction par matricule (mois le plus recent) pour ventiler les presents.
    fonctions = {}
    for mat, fonction in (
        db.session.query(Ouvrier.matricule_chantier, Ouvrier.fonction)
        .filter(Ouvrier.projet_id == pid)
        .order_by(Ouvrier.mois.desc())
        .all()
    ):
        m = (mat or "").strip()
        if m and m not in fonctions:
            fonctions[m] = (fonction or "").strip() or "Non renseigné"

    presents = [l for l in lignes if l["entree"]]
    sur_site = [l for l in presents if not l["sortie"]]
    par_fonction, par_heure = Counter(), Counter()
    for l in presents:
        par_fonction[fonctions.get((l["matricule"] or "").strip(), "Non renseigné")] += 1
        par_heure[l["entree"].hour] += 1

    base = effectif or len(presents)
    stats = {
        "effectif": effectif,
        "presents": len(presents),
        "sur_site": len(sur_site),
        "partis": len(presents) - len(sur_site),
        "absents": max(0, effectif - len(presents)),
        "taux": round(100.0 * len(presents) / base, 1) if base else 0.0,
        "par_fonction": sorted(par_fonction.items(), key=lambda x: -x[1])[:8],
        "par_heure": [[f"{h:02d}h", par_heure.get(h, 0)] for h in range(6, 20)],
    }

    return render_template(
        "presences.html",
        page="presences",
        jour=jour,
        lignes=lignes,
        stats=stats,
        total_ouvriers=effectif,
        aujourdhui=date.today(),
    )


# --------------------------------------------------------------------------
@bp.route("/api/checkin", methods=["POST"])
@login_required
@exige("pointage")
def checkin():
    d = request.get_json(silent=True) or {}
    type_ = d.get("type") if d.get("type") in ("entree", "sortie") else "entree"

    pid, mat, nom, methode, erreur = _resoudre(d)
    if erreur:
        return jsonify({"ok": False, "erreur": erreur, "definitif": True}), 400

    enreg, deja = _enregistrer(
        pid, mat, nom, type_, methode, current_user.username,
        heure=heure_appareil(d.get("heure")),
    )
    presents = Presence.query.filter_by(projet_id=pid, jour=date.today(), type="entree").count()
    return jsonify({
        "ok": True,
        "nom": nom or mat or "Ouvrier",
        "matricule": mat,
        "fonction": _fonction(pid, mat, nom),
        "type": type_,
        "heure": enreg.heure.strftime("%H:%M"),
        "deja": deja,
        "presents": presents,
    })


# --------------------------------------------------------------------------
@bp.route("/api/checkin/lot", methods=["POST"])
@login_required
@exige("pointage")
def checkin_lot():
    """Vide la file d'attente d'un telephone qui a pointe hors ligne.

    Chaque element porte l'heure relevee par l'appareil et un identifiant
    client (`uuid`) qui sert uniquement au telephone a retirer les bonnes
    lignes de sa file. L'idempotence, elle, est assuree par la contrainte
    d'unicite (projet, jour, matricule, nom, type) du modele Presence :
    rejouer deux fois le meme lot ne cree pas de doublon.

    Les erreurs de validation sont marquees `definitif` : le client doit alors
    abandonner la ligne au lieu de la reessayer sans fin.
    """
    d = request.get_json(silent=True) or {}
    items = d.get("pointages")
    if not isinstance(items, list):
        return jsonify({"ok": False, "erreur": "Liste de pointages attendue."}), 400
    if len(items) > MAX_LOT:
        return jsonify({"ok": False, "erreur": f"Lot trop volumineux (max {MAX_LOT})."}), 413

    resultats, projets = [], set()
    for item in items:
        if not isinstance(item, dict):
            continue
        uid = item.get("uuid")
        type_ = item.get("type") if item.get("type") in ("entree", "sortie") else "entree"
        pid, mat, nom, methode, erreur = _resoudre(item)
        if erreur:
            resultats.append({"uuid": uid, "ok": False, "erreur": erreur, "definitif": True})
            continue
        try:
            _enregistrer(
                pid, mat, nom, type_, methode, current_user.username,
                heure=heure_appareil(item.get("heure")), differe=True, commit=False,
            )
            db.session.commit()
        except IntegrityError:
            # Course entre deux telephones sur le meme badge : la contrainte
            # d'unicite a tranche, la presence existe donc bien. Rien a refaire.
            db.session.rollback()
        except SQLAlchemyError:
            db.session.rollback()
            resultats.append({"uuid": uid, "ok": False, "erreur": "Enregistrement impossible."})
            continue
        projets.add(pid)
        resultats.append({"uuid": uid, "ok": True})

    pid_actif = projet_actif_id()
    presents = Presence.query.filter_by(
        projet_id=pid_actif, jour=date.today(), type="entree"
    ).count()
    return jsonify({
        "ok": True,
        "resultats": resultats,
        "recus": sum(1 for r in resultats if r["ok"]),
        "presents": presents,
    })


# --------------------------------------------------------------------------
# Auto-pointage : l'ouvrier scanne son propre badge avec son telephone.
# Page publique (le badge signe fait foi) — aucune donnee sensible affichee.
# --------------------------------------------------------------------------
@bp.route("/p/<token>", methods=["GET", "POST"])
def self_checkin(token):
    info = svc_badges.decoder_token(token)
    if info is None:
        return render_template("self_checkin.html", erreur="Badge non reconnu."), 400
    pid, mat, nom = info
    projet = db.session.get(Projet, pid)
    pin_requis = (projet.pin_pointage or "").strip() if projet else ""

    sens = request.values.get("s")
    type_ = sens if sens in ("entree", "sortie") else "entree"

    # Si un code de pointage est defini pour le projet, l'ouvrier doit le saisir.
    fourni = (request.values.get("pin") or "").strip()
    if pin_requis and fourni != pin_requis:
        return render_template(
            "self_checkin.html",
            demande_pin=True,
            token=token,
            projet=projet,
            nom=nom or mat or "Ouvrier",
            pin_errone=bool(fourni),
        )

    enreg, deja = _enregistrer(pid, mat, nom, type_, methode="self", saisi_par="auto")
    return render_template(
        "self_checkin.html",
        nom=nom or mat or "Ouvrier",
        matricule=mat,
        type=type_,
        heure=enreg.heure.strftime("%H:%M"),
        deja=deja,
        projet=projet,
        token=token,
        pin=fourni if pin_requis else "",
    )


# --------------------------------------------------------------------------
# Export des presences du jour (Excel / CSV)
# --------------------------------------------------------------------------
@bp.route("/presences/export.<fmt>")
@login_required
@exige("pointage")
def export_presences(fmt):
    pid = projet_actif_id()
    valeur = request.args.get("jour")
    try:
        jour = date.fromisoformat(valeur) if valeur else date.today()
    except ValueError:
        jour = date.today()
    lignes = _presences_jour(pid, jour)

    entetes = ["Matricule", "Nom & Prenom", "Entree", "Sortie", "Duree", "Statut"]

    def cellules(l):
        duree, statut = "", "Absent"
        if l["entree"] and l["sortie"]:
            mins = int((l["sortie"] - l["entree"]).total_seconds() // 60)
            duree = f"{mins // 60}h{mins % 60:02d}"
            statut = "Parti"
        elif l["entree"]:
            statut = "Sur site"
        return [
            l["matricule"] or "", l["nom"] or "",
            l["entree"].strftime("%H:%M") if l["entree"] else "",
            l["sortie"].strftime("%H:%M") if l["sortie"] else "",
            duree, statut,
        ]

    nom_fichier = f"presences_{jour.isoformat()}.{fmt}"

    if fmt == "csv":
        tampon = io.StringIO()
        ecrivain = csv.writer(tampon, delimiter=";")
        ecrivain.writerow(entetes)
        for l in lignes:
            ecrivain.writerow(cellules(l))
        contenu = tampon.getvalue().encode("utf-8-sig")
        mime = "text/csv; charset=utf-8"
    elif fmt in ("xlsx", "excel"):
        wb = Workbook()
        ws = wb.active
        ws.title = f"Presences {jour.isoformat()}"[:31]
        ws.append(entetes)
        fill = PatternFill("solid", fgColor="1B6B43")
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = fill
            c.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for l in lignes:
            ws.append(cellules(l))
        for i, larg in enumerate((14, 30, 10, 10, 10, 12), start=1):
            ws.column_dimensions[chr(64 + i)].width = larg
        flux = io.BytesIO()
        wb.save(flux)
        contenu = flux.getvalue()
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        nom_fichier = f"presences_{jour.isoformat()}.xlsx"
    else:
        return jsonify({"erreur": "format inconnu"}), 400

    return Response(
        contenu, mimetype=mime,
        headers={"Content-Disposition": f'attachment; filename="{nom_fichier}"'},
    )
